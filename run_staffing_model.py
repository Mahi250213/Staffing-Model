import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import argparse


def run_staffing_model(input_excel, output_excel):

    def is_empty_day(df, d):
        required = ["total_rooms", "trainees", "crnas", "faculty"]
        for r in required:
            if r not in df.index:
                return True
            if pd.isna(df.loc[r, d]):
                return True
        return False

    def round_half_up(x):
        """
        Custom rounding:
        - decimal >= 0.5 → round up
        - decimal < 0.5 → round down
        """
        if x == "" or x is None:
            return ""
        return math.floor(x + 0.5)

    def round_1_decimal(x):
        if x == "" or x is None:
            return ""
        return round(x, 1)


    df = pd.read_excel(input_excel, index_col=0)
    df.index = df.index.str.strip().str.lower()
    dates = list(df.columns)

    # ---------------- Excel styles ----------------
    yellow = PatternFill("solid", fgColor="FFF200")
    blue = PatternFill("solid", fgColor="CFEAF7")
    green = PatternFill("solid", fgColor="E3F4D7")
    bold = Font(bold=True)

    out_wb = load_workbook(input_excel)
    ws = out_wb.active

    # Clear worksheet
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)

    # ---------------- Column layout ----------------
    col_positions = {}
    col = 3
    for i, d in enumerate(dates):
        col_positions[d] = col
        col += 1
        if (i + 1) % 5 == 0:
            col += 1  # weekly gap

    # ---------------- Headers ----------------
    for d, c in col_positions.items():
        ws.cell(row=1, column=c).value = d.strftime("%d-%b")
        ws.cell(row=1, column=c).font = bold

    ws["B2"] = "Main/NL/ASC"
    ws["B2"].font = bold

    for d, c in col_positions.items():
        ws.cell(row=2, column=c).value = d.strftime("%a")
        ws.cell(row=2, column=c).font = bold

    row = 3

    def write_row(label, values, fill=None, bold_row=False):
        nonlocal row
    
        header_cell = ws[f"B{row}"]
        header_cell.value = label
    
        if bold_row:
            header_cell.font = bold
    
        # 🔹 NEW: apply same fill to header cell
        if fill:
            header_cell.fill = fill
    
        for d, val in values.items():
            cell = ws.cell(row=row, column=col_positions[d])
            cell.value = val
            if fill:
                cell.fill = fill
                
        row += 1

    def blank():
        nonlocal row
        row += 1

    computed = {}

    print("\n================ STAFFING MODEL RUN START ================\n")

    for d in dates:
        print("\n----------------------------------------------------------")
        print(f"Date: {d}")

        # ---- EMPTY INPUT HANDLING ----
        if is_empty_day(df, d):
            print("  ⚠ Input data is empty for this date. Output will remain blank.")
    
            computed[d] = {
                "nfp": "",
                "no_flex": "",
                "trainee": "",
                "solo": "",
                "crna": "",
                "diff": "",
                "1:1": "",
                "1:2": "",
                "1:3.5": "",
                "supervisory": "",
                "addition_supervisory_faculty": "",
                "faculty_needed": "",
                "final_faculty": "",
                "nl": "",
                "mor": "",
                "pct_solo": "",
                "faculty_sched": "",
                "overage": "",
                "crna_sched": "",
                "crna_demand": "",
                "crna_needed": "",
            }
            continue

        total_rooms = int(df.loc["total_rooms", d])
        trainees = int(df.loc["trainees", d])
        crnas = int(df.loc["crnas", d])
        faculty = int(df.loc["faculty", d])

        fixed_crnas = 6
        cardiac_faculty = 3
        nl_fac = 5

        scheduled_flex_crnas = max(crnas - fixed_crnas, 0)

        print("\nINPUTS")
        print(f"  Rooms      : {total_rooms}")
        print(f"  Trainees   : {trainees}")
        print(f"  CRNAs      : {crnas}")
        print(f"  Faculty    : {faculty}")

        # -------- Initial availability --------
        available_rooms = total_rooms - fixed_crnas - cardiac_faculty
        available_faculty = faculty - cardiac_faculty

        print("\nINITIAL AVAILABILITY")
        print(
            f"  Available rooms after fixed & cardiac : "
            f"{total_rooms} - {fixed_crnas} - {cardiac_faculty} = {available_rooms}"
        )
        print(
            f"  Available faculty after cardiac       : "
            f"{faculty} - {cardiac_faculty} = {available_faculty}"
        )

        # -------- Trainee coverage --------
        trainee_rooms = min(trainees, available_rooms)
        raw_trainee_faculty = trainee_rooms / 2 - 1.5
        faculty_for_trainees = round_1_decimal(min(
            raw_trainee_faculty,
            available_faculty
        ))

        prev_rooms = available_rooms
        prev_fac = available_faculty

        available_rooms -= trainee_rooms
        available_faculty -= faculty_for_trainees

        print("\nTRAINEE COVERAGE (1:2)")
        print(f"  Trainee rooms assigned : min({trainees}, {prev_rooms}) = {trainee_rooms}")
        print(f"  Faculty supervising    : ({trainee_rooms} / 2) - 1.5 = {faculty_for_trainees}")
        print(f"  Rooms remaining        : {prev_rooms} - {trainee_rooms} = {available_rooms}")
        print(
            f"  Faculty remaining      : "
            f"{prev_fac} - {faculty_for_trainees} = {available_faculty}"
        )

        # -------- CRNA coverage --------
        prev_rooms = available_rooms
        prev_fac = available_faculty

        crna_rooms = min(scheduled_flex_crnas, available_rooms)
        raw_crna_faculty = crna_rooms / 3.5
        faculty_for_crnas = round_1_decimal(min(
            raw_crna_faculty,
            available_faculty
        ))

        available_rooms -= crna_rooms
        available_faculty -= faculty_for_crnas

        print("\nCRNA COVERAGE (1:3.5)")
        print(
            f"  CRNA rooms assigned    : "
            f"min({scheduled_flex_crnas}, {prev_rooms}) = {crna_rooms}"
        )
        print(
            f"  Faculty supervising    : "
            f"({crna_rooms} / 3.5) = {faculty_for_crnas}"
        )
        print(f"  Rooms remaining        : {prev_rooms} - {crna_rooms} = {available_rooms}")
        print(
            f"  Faculty remaining      : "
            f"{prev_fac} - {faculty_for_crnas} = {available_faculty}"
        )

        # -------- Solo faculty --------
        solo_faculty = min(available_faculty, available_rooms)

        print("\nSOLO FACULTY COVERAGE")
        print(
            f"  Solo faculty rooms     : "
            f"min({available_faculty}, {available_rooms}) = {solo_faculty}"
        )

        # -------- Faculty summary --------
        supervisory = cardiac_faculty + faculty_for_trainees + faculty_for_crnas
        addition_faculty_supervisory = supervisory + solo_faculty
        faculty_needed = round_half_up(supervisory + solo_faculty)
        final_faculty_required = faculty_needed + 2
        overage_faculty = faculty - final_faculty_required

        mor_asc_sat = final_faculty_required - nl_fac

        percent_solo = (
            round((solo_faculty / mor_asc_sat) * 100)
            if mor_asc_sat > 0 else 0
        )

        print("\nFACULTY SUMMARY")
        print(
            f"  Supervisory faculty    : "
            f"{cardiac_faculty} + {faculty_for_trainees} + {faculty_for_crnas} = {supervisory}"
        )
        print(f"  Solo faculty           : {solo_faculty}")
        print(
            f"  Faculty covering rooms : "
            f"{supervisory} + {solo_faculty} = {faculty_needed}"
        )
        print(
            f"  Final faculty required : "
            f"{faculty_needed} + 2 (CD + NL) = {final_faculty_required}"
        )
        print(
            f"  Faculty scheduled      : "
            f"{faculty}"
        )
        print(
            f"  Overage of faculty     : "
            f"{faculty} - {final_faculty_required} = {overage_faculty}"
        )
        print(
            f"  MOR/ASC/Sat faculty    : "
            f"{final_faculty_required} - {nl_fac} = {mor_asc_sat}"
        )
        print(
            f"  % Solo                 : "
            f"({solo_faculty} / {mor_asc_sat}) * 100 = {percent_solo}%"
        )

        # -------- CRNA summary --------
        crna_demand = total_rooms - fixed_crnas - trainee_rooms - solo_faculty
        crna_needed = crna_demand - scheduled_flex_crnas

        print("\nCRNA SUMMARY")
        print(
            f"  CRNA demand            : "
            f"{total_rooms} - {fixed_crnas} - {trainee_rooms} - {solo_faculty} = {crna_demand}"
        )
        print(
            f"  CRNAs scheduled (flex) : "
            f"{crnas} - {fixed_crnas} = {scheduled_flex_crnas}"
        )
        print(
            f"  CRNAs needed           : "
            f"{crna_demand} - {scheduled_flex_crnas} = {crna_needed}"
        )

        computed[d] = {
            "nfp": total_rooms,
            "no_flex": total_rooms - fixed_crnas,
            "trainee": trainees,
            "solo": solo_faculty,
            "crna": crna_rooms,
            "diff": (total_rooms - fixed_crnas) - trainees - crna_rooms,
            "1:1": cardiac_faculty,
            "1:2": faculty_for_trainees,
            "1:3.5": faculty_for_crnas,
            "supervisory": supervisory,
            "addition_supervisory_faculty" : addition_faculty_supervisory,
            "faculty_needed": faculty_needed,
            "final_faculty": final_faculty_required,
            "nl": nl_fac,
            "mor": mor_asc_sat,
            "pct_solo": f"{percent_solo}%",
            "faculty_sched": faculty,
            "overage": faculty - final_faculty_required,
            "crna_sched": scheduled_flex_crnas,
            "crna_demand": crna_demand,
            "crna_needed": crna_needed,
        }

    print("\n================ STAFFING MODEL RUN END ==================\n")

    # -------- Excel output (unchanged) --------
    write_row("NFP demand", {d: computed[d]["nfp"] for d in dates})
    write_row("Demand (no Flex CRNAs)", {d: computed[d]["no_flex"] for d in dates})
    write_row("Main/NL/ASC Trainee", {d: computed[d]["trainee"] for d in dates})
    write_row("Solo Faculty", {d: computed[d]["solo"] for d in dates}, fill=yellow)
    write_row("Main/NL/ASC CRNA", {d: computed[d]["crna"] for d in dates})
    write_row("difference", {d: computed[d]["diff"] for d in dates}, fill=blue)

    blank()
    ws[f"B{row}"] = "Room ratio"
    ws[f"B{row}"].font = bold
    row += 1

    write_row("1:1", {d: computed[d]["1:1"] for d in dates})
    write_row("1:2", {d: computed[d]["1:2"] for d in dates})
    write_row("1:3.5", {d: computed[d]["1:3.5"] for d in dates})

    blank()
    write_row("Supervisory Faculty needed", {d: computed[d]["supervisory"] for d in dates})
    write_row("Solo Faculty", {d: computed[d]["solo"] for d in dates})
    # write_row("", {d: computed[d]["addition_supervisory_faculty"] for d in dates})
    write_row(
        "",
        {
            d: computed[d]["addition_supervisory_faculty"]
            if isinstance(computed[d]["addition_supervisory_faculty"], (int, float))
            else ""
            for d in dates
        },
        bold_row=True
    )
    
    blank()
    write_row(
        "Faculty needed",
        {
            d: computed[d]["faculty_needed"]
            if isinstance(computed[d]["faculty_needed"], (int, float))
            else ""
            for d in dates
        },
        bold_row=True
    )


    # write_row("+ CD (MOR)", {d: 1 for d in dates})
    # write_row("+ NL OR Block", {d: 1 for d in dates})
    write_row(
        "+ CD (MOR)",
        {d: 1 if computed[d]["final_faculty"] != "" else "" for d in dates}
    )
    
    write_row(
        "+ NL OR Block",
        {d: 1 if computed[d]["final_faculty"] != "" else "" for d in dates}
    )

    write_row("", {d: computed[d]["final_faculty"] for d in dates})

    blank()
    write_row("NL fac", {d: computed[d]["nl"] for d in dates}, bold_row=True)
    write_row("MOR/ASC/Sat", {d: computed[d]["mor"] for d in dates}, bold_row=True)

    blank()
    write_row("% solo", {d: computed[d]["pct_solo"] for d in dates}, fill=green)

    blank()
    write_row("Faculty Scheduled", {d: computed[d]["faculty_sched"] for d in dates})
    write_row("Expec. Solo Faculty", {d: computed[d]["solo"] for d in dates})
    write_row("Overage of Faculty:", {d: computed[d]["overage"] for d in dates}, fill=blue)

    blank()
    write_row("CRNAs Scheduled:", {d: computed[d]["crna_sched"] for d in dates})
    write_row("CRNA Demand:", {d: computed[d]["crna_demand"] for d in dates})
    write_row("CRNAs needed:", {d: computed[d]["crna_needed"] for d in dates}, fill=yellow)

    for c in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    out_wb.save(output_excel)
    print(f"Final formatted staffing report written to: {output_excel}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Input Excel file")
    args = parser.parse_args()

    output_file = args.input.replace(".xlsx", "_output.xlsx")
    run_staffing_model(args.input, output_file)
